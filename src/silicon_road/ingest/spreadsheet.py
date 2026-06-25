"""
Load component_inventory.xlsx and convert each component to a text representation
suitable for embedding.

Spreadsheet structure:
  Sheets: ICs, Diodes, Resistors, MOSFETs, Transistors, Capacitors, Inductors, Misc
  Columns: Part Number, Manufacturer, Description, Package, Function/Category,
           Supply Voltage (V), Pincount, Datasheet URL, Qty, Location/Bin, Notes
"""

from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Column names as they appear in the spreadsheet (case-insensitive match used below)
EXPECTED_COLUMNS = [
    "Part Number",
    "Manufacturer",
    "Description",
    "Package",
    "Function/Category",
    "Supply Voltage (V)",
    "Pincount",
    "Datasheet URL",
    "Qty",
    "Location/Bin",
    "Notes",
]


@dataclass
class Component:
    """A single component from the inventory."""

    sheet: str                          # Which sheet/category it came from
    part_number: str
    manufacturer: str = ""
    description: str = ""
    package: str = ""
    category: str = ""
    supply_voltage: str = ""
    pincount: str = ""
    datasheet_url: str = ""
    qty: str = ""
    location: str = ""
    notes: str = ""
    raw: dict[str, Any] = field(default_factory=dict)
    # Set by load_inventory() when the same part number appears more than once
    # in the same sheet. First occurrence = 0 (no suffix), subsequent = 2, 3, …
    _seq: int = field(default=0, compare=False)

    @property
    def doc_id(self) -> str:
        suffix = f"_{self._seq}" if self._seq else ""
        slug = f"{self.sheet}_{self.part_number}{suffix}".lower()
        # Replace non-alphanumeric with underscores
        return "".join(c if c.isalnum() else "_" for c in slug).strip("_")

    def to_text(self) -> str:
        """
        Build a human-readable text representation for embedding.
        Richer text → better semantic retrieval.
        """
        parts = []

        parts.append(f"Category: {self.sheet}")
        if self.part_number:
            parts.append(f"Part number: {self.part_number}")
        if self.manufacturer:
            parts.append(f"Manufacturer: {self.manufacturer}")
        if self.description:
            parts.append(f"Description: {self.description}")
        if self.category:
            parts.append(f"Function: {self.category}")
        if self.package:
            parts.append(f"Package: {self.package}")
        if self.supply_voltage:
            parts.append(f"Supply voltage: {self.supply_voltage} V")
        if self.pincount:
            parts.append(f"Pin count: {self.pincount}")
        if self.qty:
            parts.append(f"Quantity on hand: {self.qty}")
        if self.location:
            parts.append(f"Bin / location: {self.location}")
        if self.notes:
            parts.append(f"Notes: {self.notes}")
        # Deliberately omit datasheet_url from the embedding text —
        # it's metadata, not semantics.

        return ". ".join(parts) + "."

    def to_metadata(self) -> dict[str, str]:
        """ChromaDB metadata dict — only string values."""
        return {
            "sheet": self.sheet,
            "part_number": self.part_number,
            "manufacturer": self.manufacturer,
            "description": self.description,
            "package": self.package,
            "category": self.category,
            "supply_voltage": str(self.supply_voltage),
            "pincount": str(self.pincount),
            "datasheet_url": self.datasheet_url,
            "qty": str(self.qty),
            "location": self.location,
            "notes": self.notes,
        }


def _cell_str(cell) -> str:
    """Return a cell's value as a clean string, or empty string if None."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _map_headers(row) -> dict[str, int]:
    """Return a mapping of canonical column name → zero-based column index."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        val = _cell_str(cell).lower()
        for col in EXPECTED_COLUMNS:
            if col.lower() == val:
                mapping[col] = idx
                break
    return mapping


def load_inventory(xlsx_path: str | Path) -> list[Component]:
    """
    Load all sheets from the inventory spreadsheet and return a flat list of
    Component objects, skipping rows that have no Part Number.

    Duplicate part numbers within the same sheet get a _seq suffix (2, 3, …)
    so every doc_id is globally unique.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Inventory file not found: {xlsx_path}")

    logger.info("Loading inventory from %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    components: list[Component] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows())
        if not rows:
            logger.warning("Sheet %r is empty, skipping", sheet_name)
            continue

        # First row is the header
        headers = _map_headers(rows[0])
        if not headers:
            logger.warning("Sheet %r has no recognizable headers, skipping", sheet_name)
            continue

        def get(row, col_name: str) -> str:
            idx = headers.get(col_name)
            if idx is None:
                return ""
            try:
                return _cell_str(row[idx])
            except IndexError:
                return ""

        # Track how many times each part number has appeared in this sheet
        seen: dict[str, int] = defaultdict(int)

        sheet_count = 0
        for row in rows[1:]:
            part_number = get(row, "Part Number")
            if not part_number:
                continue  # blank row

            seen[part_number] += 1
            seq = seen[part_number] if seen[part_number] > 1 else 0

            comp = Component(
                sheet=sheet_name,
                part_number=part_number,
                manufacturer=get(row, "Manufacturer"),
                description=get(row, "Description"),
                package=get(row, "Package"),
                category=get(row, "Function/Category"),
                supply_voltage=get(row, "Supply Voltage (V)"),
                pincount=get(row, "Pincount"),
                datasheet_url=get(row, "Datasheet URL"),
                qty=get(row, "Qty"),
                location=get(row, "Location/Bin"),
                notes=get(row, "Notes"),
                _seq=seq,
            )
            components.append(comp)
            sheet_count += 1

        logger.info("  %-14s → %d components", sheet_name, sheet_count)

    logger.info("Total: %d components loaded", len(components))
    return components
